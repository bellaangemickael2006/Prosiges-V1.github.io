"""
Génération des livrables : classeur Excel (journal + comptes) et rapport PDF
d'analyse destiné aux entreprises accompagnées.

Les deux fonctions renvoient un objet BytesIO prêt à être envoyé via
`flask.send_file`, sans écriture sur disque.
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate, Spacer,
                                Table, TableStyle)

import comptes
from models import Departement, ModePaiement, TypeOperation

BLEU = '1C3D5A'
VERT = '2E8B57'
ROUGE = 'C0392B'
GRIS = 'F0F2F5'

FORMAT_MONTANT = '#,##0 "F"'


def formater_montant(valeur):
    return f"{valeur:,.0f} FCFA".replace(',', ' ')


# ==========================================================================
# EXPORT EXCEL
# ==========================================================================

def _style_entete(cellule):
    cellule.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    cellule.fill = PatternFill('solid', fgColor=BLEU)
    cellule.alignment = Alignment(horizontal='center', vertical='center')
    cellule.border = Border(bottom=Side(style='thin', color='999999'))


def _ajuster_colonnes(feuille, largeurs):
    for index, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = largeur


def _titre_feuille(feuille, titre, sous_titre=None):
    feuille['A1'] = titre
    feuille['A1'].font = Font(name='Arial', bold=True, size=14, color=BLEU)
    if sous_titre:
        feuille['A2'] = sous_titre
        feuille['A2'].font = Font(name='Arial', italic=True, size=10)


def _feuille_journal_de_bord(classeur, titre_perimetre, annee, scope):
    """Journal de bord : entrées, sorties et solde cumulé en formules.

    Le solde de chaque ligne est une formule qui reprend celui de la ligne
    précédente : si un montant est corrigé dans le classeur, toute la
    colonne se recalcule d'elle-même.
    """
    feuille = classeur.active
    feuille.title = 'Journal de bord'
    _titre_feuille(feuille, 'JOURNAL DE BORD',
                   f"{titre_perimetre} — Exercice {annee}")

    entetes = ['Date', 'Libellé', 'Catégorie', 'Entrée', 'Sortie',
               'Mode de paiement', 'Solde', 'Pièce justificative']
    ligne_entete = 4
    for col, entete in enumerate(entetes, start=1):
        _style_entete(feuille.cell(row=ligne_entete, column=col, value=entete))

    lignes = comptes.journal_de_bord(annee, **scope)
    ligne = ligne_entete + 1
    premiere = ligne
    for element in lignes:
        feuille.cell(row=ligne, column=1,
                     value=element['date']).number_format = 'DD/MM/YYYY'
        feuille.cell(row=ligne, column=2, value=element['libelle'])
        feuille.cell(row=ligne, column=3, value=element['categorie'])
        feuille.cell(row=ligne, column=4,
                     value=element['entree'] or None).number_format = FORMAT_MONTANT
        feuille.cell(row=ligne, column=5,
                     value=element['sortie'] or None).number_format = FORMAT_MONTANT
        feuille.cell(row=ligne, column=6, value=element['mode_paiement'])

        # Solde cumulé : première ligne = entrée − sortie, puis report.
        if ligne == premiere:
            formule = f'=D{ligne}-E{ligne}'
        else:
            formule = f'=G{ligne - 1}+D{ligne}-E{ligne}'
        cellule = feuille.cell(row=ligne, column=7, value=formule)
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', size=10, bold=True)

        feuille.cell(row=ligne, column=8, value=element['preuve'])
        for col in range(1, 9):
            if col != 7:
                feuille.cell(row=ligne, column=col).font = Font(name='Arial', size=10)
        ligne += 1

    derniere = ligne - 1
    if lignes:
        feuille.cell(row=ligne, column=3, value='TOTAL').font = Font(
            name='Arial', bold=True, size=11)
        for col, lettre in ((4, 'D'), (5, 'E')):
            cellule = feuille.cell(row=ligne, column=col,
                                   value=f'=SUM({lettre}{premiere}:{lettre}{derniere})')
            cellule.number_format = FORMAT_MONTANT
            cellule.font = Font(name='Arial', bold=True, size=11)
        cellule = feuille.cell(row=ligne, column=7, value=f'=G{derniere}')
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', bold=True, size=11, color=BLEU)

    _ajuster_colonnes(feuille, [12, 34, 30, 16, 16, 18, 16, 22])
    feuille.freeze_panes = f'A{ligne_entete + 1}'
    return feuille


# --- Comptes détaillés : trois colonnes par mois -------------------------

def _colonne_mois(mois, decalage):
    """Colonne Excel d'un mois donné. `decalage` : 0 prévu, 1 réalisé, 2 écart."""
    return 2 + (mois - 1) * 3 + decalage


COLONNE_TOTAL = 2 + 12 * 3          # première colonne du total annuel (38 = AL)


def _entetes_compte(feuille, ligne):
    """Deux lignes d'en-tête : les mois, puis Prévu / Réalisé / Écart."""
    _style_entete(feuille.cell(row=ligne, column=1, value='LIBELLE'))
    _style_entete(feuille.cell(row=ligne + 1, column=1, value=''))

    for mois in range(1, 13):
        depart = _colonne_mois(mois, 0)
        cellule = feuille.cell(row=ligne, column=depart,
                               value=comptes.MOIS_FR[f'{mois:02d}'])
        _style_entete(cellule)
        feuille.merge_cells(start_row=ligne, start_column=depart,
                            end_row=ligne, end_column=depart + 2)
        for index, libelle in enumerate(('Prévu', 'Réalisé', 'Écart')):
            _style_entete(feuille.cell(row=ligne + 1, column=depart + index,
                                       value=libelle))

    _style_entete(feuille.cell(row=ligne, column=COLONNE_TOTAL,
                               value='TOTAL ANNÉE'))
    feuille.merge_cells(start_row=ligne, start_column=COLONNE_TOTAL,
                        end_row=ligne, end_column=COLONNE_TOTAL + 2)
    for index, libelle in enumerate(('Prévu', 'Réalisé', 'Écart')):
        _style_entete(feuille.cell(row=ligne + 1, column=COLONNE_TOTAL + index,
                                   value=libelle))


def _ecrire_ligne_compte(feuille, ligne_excel, donnees, gras=False,
                         fond=None, formules_totaux=True):
    """Écrit une rubrique : Prévu et Réalisé en valeurs, Écart en formule.

    L'écart et les totaux annuels sont des formules Excel : le classeur
    exporté reste juste si l'utilisateur retouche un montant.
    """
    police = Font(name='Arial', size=10, bold=gras)
    cellule = feuille.cell(row=ligne_excel, column=1, value=donnees['libelle'])
    cellule.font = Font(name='Arial', size=10, bold=gras)
    cellule.alignment = Alignment(wrap_text=True, vertical='center')

    for mois in range(1, 13):
        depart = _colonne_mois(mois, 0)
        prevu = feuille.cell(row=ligne_excel, column=depart,
                             value=round(donnees['prevu'][mois - 1]) or None)
        realise = feuille.cell(row=ligne_excel, column=depart + 1,
                               value=round(donnees['realise'][mois - 1]) or None)
        lettre_p = get_column_letter(depart)
        lettre_r = get_column_letter(depart + 1)
        ecart = feuille.cell(row=ligne_excel, column=depart + 2,
                             value=f'={lettre_r}{ligne_excel}-{lettre_p}{ligne_excel}')
        for cellule in (prevu, realise, ecart):
            cellule.number_format = FORMAT_MONTANT
            cellule.font = police

    for index in range(3):
        colonne = COLONNE_TOTAL + index
        if formules_totaux:
            # Somme des douze colonnes de même nature (une sur trois)
            references = ','.join(
                f'{get_column_letter(_colonne_mois(m, index))}{ligne_excel}'
                for m in range(1, 13))
            valeur = f'=SUM({references})'
        else:
            # Un solde ne se cumule pas : c'est celui de décembre.
            lettre = get_column_letter(_colonne_mois(12, index))
            valeur = f'={lettre}{ligne_excel}'
        cellule = feuille.cell(row=ligne_excel, column=colonne, value=valeur)
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', size=10, bold=True)

    if fond:
        for colonne in range(1, COLONNE_TOTAL + 3):
            feuille.cell(row=ligne_excel, column=colonne).fill = PatternFill(
                'solid', fgColor=fond)


def _somme_colonnes(feuille, ligne_excel, premiere, derniere, libelle,
                    fond=GRIS):
    """Ligne de total d'une section, en formules SUM sur chaque colonne."""
    cellule = feuille.cell(row=ligne_excel, column=1, value=libelle)
    cellule.font = Font(name='Arial', size=10, bold=True)
    for colonne in range(2, COLONNE_TOTAL + 3):
        lettre = get_column_letter(colonne)
        cellule = feuille.cell(row=ligne_excel, column=colonne,
                               value=f'=SUM({lettre}{premiere}:{lettre}{derniere})')
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', size=10, bold=True)
    for colonne in range(1, COLONNE_TOTAL + 3):
        feuille.cell(row=ligne_excel, column=colonne).fill = PatternFill(
            'solid', fgColor=fond)


def _difference_lignes(feuille, ligne_excel, ligne_gauche, ligne_droite,
                       libelle, couleur=BLEU):
    """Ligne calculée : gauche − droite, colonne par colonne."""
    cellule = feuille.cell(row=ligne_excel, column=1, value=libelle)
    cellule.font = Font(name='Arial', size=10, bold=True, color=couleur)
    for colonne in range(2, COLONNE_TOTAL + 3):
        lettre = get_column_letter(colonne)
        cellule = feuille.cell(
            row=ligne_excel, column=colonne,
            value=f'={lettre}{ligne_gauche}-{lettre}{ligne_droite}')
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', size=10, bold=True, color=couleur)
    for colonne in range(1, COLONNE_TOTAL + 3):
        feuille.cell(row=ligne_excel, column=colonne).fill = PatternFill(
            'solid', fgColor='E8EDF3')


def _mise_en_page_compte(feuille):
    feuille.column_dimensions['A'].width = 46
    for colonne in range(2, COLONNE_TOTAL + 3):
        feuille.column_dimensions[get_column_letter(colonne)].width = 13
    feuille.freeze_panes = 'B6'


def _feuille_exploitation(classeur, compte, titre_perimetre, annee):
    """Compte d'exploitation au format du cabinet."""
    feuille = classeur.create_sheet("Compte d'exploitation")
    _titre_feuille(feuille, "COMPTE D'EXPLOITATION",
                   f"{titre_perimetre} — Exercice {annee}")
    _entetes_compte(feuille, 4)

    ligne = 6
    totaux = {}
    for section in compte['sections']:
        cellule = feuille.cell(row=ligne, column=1, value=section['titre'])
        cellule.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        for colonne in range(1, COLONNE_TOTAL + 3):
            feuille.cell(row=ligne, column=colonne).fill = PatternFill(
                'solid', fgColor=BLEU)
        ligne += 1

        premiere = ligne
        for rubrique in section['lignes']:
            _ecrire_ligne_compte(feuille, ligne, rubrique)
            ligne += 1
        derniere = ligne - 1

        if derniere >= premiere:
            _somme_colonnes(feuille, ligne, premiere, derniere,
                            section['total']['libelle'])
        else:
            _ecrire_ligne_compte(feuille, ligne, section['total'], gras=True,
                                 fond=GRIS)
        totaux[section['code']] = ligne
        ligne += 1

        if section['code'] == 'variables':
            _difference_lignes(feuille, ligne, totaux['produits'],
                               totaux['variables'], 'III - MARGE BRUTE', VERT)
            totaux['marge_brute'] = ligne
            ligne += 1

    _difference_lignes(feuille, ligne, totaux['marge_brute'], totaux['fixes'],
                       "V - RÉSULTAT D'EXPLOITATION (Bénéfice ou Perte)", BLEU)
    ligne += 2
    feuille.cell(row=ligne, column=1,
                 value="Écart = Réalisé − Prévu. Le prévu vient du budget saisi "
                       "par le gérant ; le réalisé est calculé automatiquement "
                       "à partir du journal de bord.").font = Font(
        name='Arial', italic=True, size=9, color='777777')

    _mise_en_page_compte(feuille)
    return feuille


def _feuille_tresorerie(classeur, compte, titre_perimetre, annee):
    """Compte de trésorerie au format du cabinet."""
    feuille = classeur.create_sheet('Compte de trésorerie')
    _titre_feuille(feuille, 'COMPTE DE TRÉSORERIE',
                   f"{titre_perimetre} — Exercice {annee}")
    _entetes_compte(feuille, 4)

    ligne = 6
    totaux = {}
    for section in compte['sections']:
        cellule = feuille.cell(row=ligne, column=1, value=section['titre'])
        cellule.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        for colonne in range(1, COLONNE_TOTAL + 3):
            feuille.cell(row=ligne, column=colonne).fill = PatternFill(
                'solid', fgColor=BLEU)
        ligne += 1

        premiere = ligne
        for rubrique in section['lignes']:
            _ecrire_ligne_compte(feuille, ligne, rubrique)
            ligne += 1
        _somme_colonnes(feuille, ligne, premiere, ligne - 1,
                        section['total']['libelle'])
        totaux[section['code']] = ligne
        ligne += 2

    # --- Soldes : le début de période reporte la fin du mois précédent ---
    ligne_debut = ligne
    _ecrire_ligne_compte(feuille, ligne_debut, compte['solde_debut'],
                         gras=True, fond='F2F5F8', formules_totaux=False)
    ligne_fin = ligne_debut + 1
    _ecrire_ligne_compte(feuille, ligne_fin, compte['solde_fin'],
                         gras=True, fond='E8EDF3', formules_totaux=False)

    encaissements = totaux['encaissements']
    decaissements = totaux['decaissements']
    for mois in range(1, 13):
        for decalage in range(2):        # colonnes Prévu et Réalisé
            colonne = _colonne_mois(mois, decalage)
            lettre = get_column_letter(colonne)
            if mois > 1:
                precedente = get_column_letter(_colonne_mois(mois - 1, decalage))
                feuille.cell(row=ligne_debut, column=colonne,
                             value=f'={precedente}{ligne_fin}')
            feuille.cell(
                row=ligne_fin, column=colonne,
                value=f'={lettre}{ligne_debut}+{lettre}{encaissements}'
                      f'-{lettre}{decaissements}')
            for rangee in (ligne_debut, ligne_fin):
                feuille.cell(row=rangee, column=colonne).number_format = FORMAT_MONTANT
                feuille.cell(row=rangee, column=colonne).font = Font(
                    name='Arial', size=10, bold=True)

    ligne = ligne_fin + 2
    feuille.cell(row=ligne, column=1,
                 value="Solde de fin de période (D) = Solde de début (C) "
                       "+ Total des encaissements (A) − Total des "
                       "décaissements (B).").font = Font(
        name='Arial', italic=True, size=9, color='777777')

    _mise_en_page_compte(feuille)
    return feuille


def _feuille_produits(classeur, entreprise_id):
    """Liste des produits et services, et familles de produits."""
    from models import Article, FamilleProduit, NatureArticle

    feuille = classeur.create_sheet('Produits et services')
    _titre_feuille(feuille, 'LISTE DES PRODUITS ET SERVICES')

    entetes = ['Référence', 'Désignation', 'Nature', 'Famille', 'Unité',
               'Prix de vente', "Prix d'achat", 'Stock', "Seuil d'alerte",
               'Valeur du stock', 'Actif']
    for col, entete in enumerate(entetes, start=1):
        _style_entete(feuille.cell(row=3, column=col, value=entete))

    ligne = 4
    articles = (Article.query.filter_by(entreprise_id=entreprise_id)
                .order_by(Article.designation).all())
    for article in articles:
        feuille.cell(row=ligne, column=1, value=article.reference or '')
        feuille.cell(row=ligne, column=2, value=article.designation)
        feuille.cell(row=ligne, column=3,
                     value=NatureArticle.LABELS.get(article.nature, ''))
        feuille.cell(row=ligne, column=4,
                     value=article.famille.nom if article.famille else '')
        feuille.cell(row=ligne, column=5, value=article.unite or '')
        feuille.cell(row=ligne, column=6,
                     value=article.prix_vente or 0).number_format = FORMAT_MONTANT
        feuille.cell(row=ligne, column=7,
                     value=article.prix_achat or 0).number_format = FORMAT_MONTANT
        if article.suit_stock:
            feuille.cell(row=ligne, column=8, value=article.quantite_stock or 0)
            feuille.cell(row=ligne, column=9, value=article.seuil_alerte or 0)
            # Valeur du stock : formule, pour rester juste après retouche
            feuille.cell(row=ligne, column=10,
                         value=f'=H{ligne}*IF(G{ligne}>0,G{ligne},F{ligne})'
                         ).number_format = FORMAT_MONTANT
        feuille.cell(row=ligne, column=11, value='Oui' if article.actif else 'Non')
        for col in range(1, 12):
            feuille.cell(row=ligne, column=col).font = Font(name='Arial', size=10)
        ligne += 1

    familles = (FamilleProduit.query.filter_by(entreprise_id=entreprise_id)
                .order_by(FamilleProduit.ordre, FamilleProduit.nom).all())
    if familles:
        ligne += 2
        feuille.cell(row=ligne, column=1, value='FAMILLES DE PRODUITS').font = Font(
            name='Arial', bold=True, size=12, color=BLEU)
        ligne += 1
        for col, entete in enumerate(['Famille', 'Description',
                                      "Nombre d'articles"], start=1):
            _style_entete(feuille.cell(row=ligne, column=col, value=entete))
        ligne += 1
        for famille in familles:
            feuille.cell(row=ligne, column=1, value=famille.nom)
            feuille.cell(row=ligne, column=2, value=famille.description or '')
            feuille.cell(row=ligne, column=3, value=len(famille.articles))
            ligne += 1

    _ajuster_colonnes(feuille, [16, 34, 12, 20, 12, 16, 16, 12, 14, 18, 10])
    return feuille


def generer_excel(operations, titre_perimetre, annee, scope):
    """Classeur de suivi complet, à l'image du classeur Google Sheets.

    Quatre feuilles pour une entreprise accompagnée : journal de bord,
    compte d'exploitation, compte de trésorerie, produits et services.

    Les écarts, les totaux et les soldes sont des FORMULES Excel et non
    des valeurs figées : si un montant est corrigé dans le classeur, tout
    se recalcule — c'est ce qui était demandé.
    """
    classeur = Workbook()
    entreprise_id = scope.get('entreprise_id')
    cree_par_id = scope.get('cree_par_id')

    # Le journal de bord gère lui-même les bornes de l'exercice.
    scope_perimetre = {k: v for k, v in scope.items()
                       if k in ('entreprise_id', 'departement',
                                'interne_seulement', 'cree_par_id')}

    _feuille_journal_de_bord(classeur, titre_perimetre, annee, scope_perimetre)

    if entreprise_id:
        _feuille_exploitation(
            classeur,
            comptes.compte_exploitation_detaille(entreprise_id, annee,
                                                 cree_par_id=cree_par_id),
            titre_perimetre, annee)
        _feuille_tresorerie(
            classeur,
            comptes.compte_tresorerie_detaille(entreprise_id, annee,
                                               cree_par_id=cree_par_id),
            titre_perimetre, annee)
        _feuille_produits(classeur, entreprise_id)
    else:
        _feuille_synthese_interne(classeur, titre_perimetre, annee, scope)

    flux = BytesIO()
    classeur.save(flux)
    flux.seek(0)
    return flux


def _feuille_synthese_interne(classeur, titre_perimetre, annee, scope):
    """Repli pour le périmètre interne du cabinet, qui n'a ni budget ni
    catalogue : synthèse des comptes et évolution mensuelle."""

    # ---------------- Compte d'exploitation simplifié ----------------
    exploitation = comptes.compte_exploitation(**scope)
    feuille2 = classeur.create_sheet("Exploitation")

    feuille2['A1'] = "COMPTE D'EXPLOITATION"
    feuille2['A1'].font = Font(name='Arial', bold=True, size=14, color=BLEU)
    feuille2['A2'] = f"{titre_perimetre} — Exercice {annee}"
    feuille2['A2'].font = Font(name='Arial', italic=True, size=10)

    ligne = 4
    for col, entete in enumerate(['Rubrique', 'Catégorie', 'Nombre', 'Montant (FCFA)'], start=1):
        _style_entete(feuille2.cell(row=ligne, column=col, value=entete))

    ligne += 1
    debut_produits = ligne
    for produit in exploitation['produits']:
        feuille2.cell(row=ligne, column=1, value='PRODUITS')
        feuille2.cell(row=ligne, column=2, value=produit['categorie'])
        feuille2.cell(row=ligne, column=3, value=produit['nombre'])
        feuille2.cell(row=ligne, column=4, value=produit['total']).number_format = FORMAT_MONTANT
        ligne += 1
    fin_produits = ligne - 1

    feuille2.cell(row=ligne, column=2, value='TOTAL PRODUITS').font = Font(
        name='Arial', bold=True)
    cellule_ca = feuille2.cell(
        row=ligne, column=4,
        value=f'=SUM(D{debut_produits}:D{fin_produits})' if exploitation['produits'] else 0)
    cellule_ca.number_format = FORMAT_MONTANT
    cellule_ca.font = Font(name='Arial', bold=True, color=VERT)
    ligne_ca = ligne
    ligne += 2

    debut_charges = ligne
    for charge in exploitation['detail_charges']:
        feuille2.cell(row=ligne, column=1, value='CHARGES')
        feuille2.cell(row=ligne, column=2, value=charge['categorie'])
        feuille2.cell(row=ligne, column=3, value=charge['nombre'])
        feuille2.cell(row=ligne, column=4, value=charge['total']).number_format = FORMAT_MONTANT
        ligne += 1
    fin_charges = ligne - 1

    feuille2.cell(row=ligne, column=2, value='TOTAL CHARGES').font = Font(
        name='Arial', bold=True)
    cellule_ch = feuille2.cell(
        row=ligne, column=4,
        value=f'=SUM(D{debut_charges}:D{fin_charges})' if exploitation['detail_charges'] else 0)
    cellule_ch.number_format = FORMAT_MONTANT
    cellule_ch.font = Font(name='Arial', bold=True, color=ROUGE)
    ligne_charges = ligne
    ligne += 2

    feuille2.cell(row=ligne, column=2, value='RÉSULTAT NET').font = Font(
        name='Arial', bold=True, size=12)
    cellule_res = feuille2.cell(row=ligne, column=4, value=f'=D{ligne_ca}-D{ligne_charges}')
    cellule_res.number_format = FORMAT_MONTANT
    cellule_res.font = Font(name='Arial', bold=True, size=12, color=BLEU)
    cellule_res.fill = PatternFill('solid', fgColor=GRIS)

    _ajuster_colonnes(feuille2, [16, 32, 12, 20])

    # ---------------- Feuille 3 : Trésorerie ----------------
    tresorerie = comptes.compte_tresorerie(**scope)
    feuille3 = classeur.create_sheet('Trésorerie')

    feuille3['A1'] = 'COMPTE DE TRÉSORERIE'
    feuille3['A1'].font = Font(name='Arial', bold=True, size=14, color=BLEU)
    feuille3['A2'] = f"{titre_perimetre} — Exercice {annee}"
    feuille3['A2'].font = Font(name='Arial', italic=True, size=10)

    for col, entete in enumerate(['Mode de paiement', 'Solde (FCFA)'], start=1):
        _style_entete(feuille3.cell(row=4, column=col, value=entete))

    ligne = 5
    debut_soldes = ligne
    for mode in [ModePaiement.CASH, ModePaiement.MOBILE_MONEY, ModePaiement.BANQUE]:
        feuille3.cell(row=ligne, column=1, value=ModePaiement.LABELS[mode])
        feuille3.cell(row=ligne, column=2,
                      value=tresorerie[mode]).number_format = FORMAT_MONTANT
        ligne += 1
    fin_soldes = ligne - 1

    feuille3.cell(row=ligne, column=1, value='TRÉSORERIE DISPONIBLE').font = Font(
        name='Arial', bold=True)
    cellule = feuille3.cell(row=ligne, column=2, value=f'=SUM(B{debut_soldes}:B{fin_soldes})')
    cellule.number_format = FORMAT_MONTANT
    cellule.font = Font(name='Arial', bold=True, color=VERT)
    ligne += 2

    feuille3.cell(row=ligne, column=1, value='Créances à encaisser (crédit)').font = Font(
        name='Arial', italic=True)
    feuille3.cell(row=ligne, column=2,
                  value=tresorerie['creances']).number_format = FORMAT_MONTANT
    ligne += 1
    feuille3.cell(row=ligne, column=1,
                  value='Note : les ventes à crédit ne sont pas encore encaissées, '
                        'elles sont donc exclues de la trésorerie disponible.').font = Font(
        name='Arial', italic=True, size=9, color='777777')

    _ajuster_colonnes(feuille3, [46, 20])

    # ---------------- Feuille 4 : Résumé mensuel ----------------
    resume = comptes.resume_mensuel(annee, **scope)
    feuille4 = classeur.create_sheet('Résumé mensuel')

    feuille4['A1'] = f'RÉSUMÉ MENSUEL {annee}'
    feuille4['A1'].font = Font(name='Arial', bold=True, size=14, color=BLEU)
    feuille4['A2'] = titre_perimetre
    feuille4['A2'].font = Font(name='Arial', italic=True, size=10)

    for col, entete in enumerate(['Mois', 'Ventes', 'Achats', 'Résultat'], start=1):
        _style_entete(feuille4.cell(row=4, column=col, value=entete))

    ligne = 5
    debut = ligne
    for cle in sorted(resume.keys()):
        val = resume[cle]
        feuille4.cell(row=ligne, column=1, value=val['libelle'])
        feuille4.cell(row=ligne, column=2, value=val['ventes']).number_format = FORMAT_MONTANT
        feuille4.cell(row=ligne, column=3, value=val['achats']).number_format = FORMAT_MONTANT
        cellule = feuille4.cell(row=ligne, column=4, value=f'=B{ligne}-C{ligne}')
        cellule.number_format = FORMAT_MONTANT
        ligne += 1
    fin = ligne - 1

    feuille4.cell(row=ligne, column=1, value='TOTAL ANNÉE').font = Font(name='Arial', bold=True)
    for col in (2, 3, 4):
        lettre = get_column_letter(col)
        cellule = feuille4.cell(row=ligne, column=col, value=f'=SUM({lettre}{debut}:{lettre}{fin})')
        cellule.number_format = FORMAT_MONTANT
        cellule.font = Font(name='Arial', bold=True)

    _ajuster_colonnes(feuille4, [18, 18, 18, 18])
    return classeur


# ==========================================================================
# EXPORT PDF — rapport d'analyse
# ==========================================================================

def _styles_pdf():
    base = getSampleStyleSheet()
    return {
        'titre': ParagraphStyle(
            'TitreAK', parent=base['Title'], fontName='Helvetica-Bold',
            fontSize=18, textColor=colors.HexColor('#1C3D5A'), spaceAfter=4),
        'sous_titre': ParagraphStyle(
            'SousTitreAK', parent=base['Normal'], fontName='Helvetica-Oblique',
            fontSize=10, textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER, spaceAfter=18),
        'section': ParagraphStyle(
            'SectionAK', parent=base['Heading2'], fontName='Helvetica-Bold',
            fontSize=13, textColor=colors.HexColor('#1C3D5A'),
            spaceBefore=16, spaceAfter=8),
        'corps': ParagraphStyle(
            'CorpsAK', parent=base['Normal'], fontName='Helvetica',
            fontSize=10, leading=15, alignment=TA_JUSTIFY, spaceAfter=6),
        'note': ParagraphStyle(
            'NoteAK', parent=base['Normal'], fontName='Helvetica-Oblique',
            fontSize=8.5, textColor=colors.HexColor('#777777'), spaceBefore=10),
    }


def _table_style_standard():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1C3D5A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FB')]),
    ])


def generer_rapport_pdf(titre_perimetre, annee, scope, auteur=None,
                        commentaire_consultant=None, operations_recentes=None):
    """Rapport d'analyse : synthèse, exploitation, trésorerie, évolution
    mensuelle, constats automatiques et espace commentaire du consultant."""
    exploitation = comptes.compte_exploitation(**scope)
    tresorerie = comptes.compte_tresorerie(**scope)
    resume = comptes.resume_mensuel(annee, **scope)
    meilleurs_clients = comptes.top_clients(limite=5, **scope)
    constats = comptes.analyse_automatique(exploitation, tresorerie, resume)

    styles = _styles_pdf()
    flux = BytesIO()
    document = SimpleDocTemplate(
        flux, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"Rapport d'analyse — {titre_perimetre} {annee}",
        author='AK World',
    )

    contenu = []
    contenu.append(Paragraph("RAPPORT D'ANALYSE D'ACTIVITÉ", styles['titre']))
    contenu.append(Paragraph(
        f"{titre_perimetre} — Exercice {annee}<br/>"
        f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        + (f" par {auteur}" if auteur else ""),
        styles['sous_titre']))

    # --- Synthèse ---
    contenu.append(Paragraph('1. Synthèse financière', styles['section']))
    donnees = [
        ['Indicateur', 'Montant'],
        ["Chiffre d'affaires", formater_montant(exploitation['chiffre_affaires'])],
        ['Total des charges', formater_montant(exploitation['charges'])],
        ['Résultat net', formater_montant(exploitation['resultat_net'])],
        ['Taux de marge', f"{exploitation['taux_marge']:.1f} %"],
        ['Trésorerie disponible', formater_montant(tresorerie['total_disponible'])],
        ['Créances à encaisser', formater_montant(tresorerie['creances'])],
    ]
    table = Table(donnees, colWidths=[9 * cm, 7 * cm])
    style = _table_style_standard()
    style.add('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold')
    style.add('TEXTCOLOR', (1, 3), (1, 3),
              colors.HexColor('#2E8B57') if exploitation['resultat_net'] >= 0
              else colors.HexColor('#C0392B'))
    table.setStyle(style)
    contenu.append(table)

    # --- Constats ---
    contenu.append(Paragraph('2. Constats', styles['section']))
    for constat in constats:
        contenu.append(Paragraph(f"• {constat}", styles['corps']))

    # --- Détail des charges ---
    if exploitation['detail_charges']:
        contenu.append(Paragraph('3. Structure des charges', styles['section']))
        donnees = [['Poste de charge', 'Opérations', 'Montant', 'Part']]
        total_charges = exploitation['charges'] or 1
        for charge in exploitation['detail_charges'][:12]:
            donnees.append([
                charge['categorie'],
                str(charge['nombre']),
                formater_montant(charge['total']),
                f"{charge['total'] / total_charges * 100:.1f} %",
            ])
        table = Table(donnees, colWidths=[7 * cm, 2.5 * cm, 4 * cm, 2.5 * cm])
        table.setStyle(_table_style_standard())
        contenu.append(table)

    contenu.append(PageBreak())

    # --- Évolution mensuelle ---
    contenu.append(Paragraph(f'4. Évolution mensuelle {annee}', styles['section']))
    donnees = [['Mois', 'Ventes', 'Achats', 'Résultat']]
    total_v = total_a = 0.0
    for cle in sorted(resume.keys()):
        val = resume[cle]
        total_v += val['ventes']
        total_a += val['achats']
        donnees.append([
            val['libelle'],
            formater_montant(val['ventes']),
            formater_montant(val['achats']),
            formater_montant(val['resultat']),
        ])
    donnees.append(['TOTAL', formater_montant(total_v), formater_montant(total_a),
                    formater_montant(total_v - total_a)])
    table = Table(donnees, colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
    style = _table_style_standard()
    style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
    style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0F2F5'))
    table.setStyle(style)
    contenu.append(table)

    # --- Meilleurs clients ---
    if meilleurs_clients:
        contenu.append(Paragraph('5. Principaux clients', styles['section']))
        donnees = [['Client', 'Opérations', "Chiffre d'affaires"]]
        for client in meilleurs_clients:
            donnees.append([client['nom'], str(client['nombre']),
                            formater_montant(client['total'])])
        table = Table(donnees, colWidths=[8 * cm, 3.5 * cm, 4.5 * cm])
        table.setStyle(_table_style_standard())
        contenu.append(table)

    # --- Commentaire du consultant ---
    contenu.append(Paragraph('6. Analyse et orientations du cabinet', styles['section']))
    if commentaire_consultant:
        for paragraphe in commentaire_consultant.split('\n'):
            if paragraphe.strip():
                contenu.append(Paragraph(paragraphe.strip(), styles['corps']))
    else:
        contenu.append(Paragraph(
            "<i>(Espace réservé aux recommandations du consultant en charge du suivi. "
            "Il peut être renseigné au moment de la génération du rapport.)</i>",
            styles['corps']))

    contenu.append(Spacer(1, 20))
    contenu.append(Paragraph(
        "Document généré automatiquement à partir du journal d'opérations saisi dans "
        "l'application de gestion AK World. Les montants sont exprimés en francs CFA (XOF).",
        styles['note']))

    document.build(contenu)
    flux.seek(0)
    return flux
